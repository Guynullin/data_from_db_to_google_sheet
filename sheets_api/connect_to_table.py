import pickle
import os

from pathlib import Path
from openpyxl import load_workbook
from .process_xlsx import process_xlsx
from .downloal_xlsx import download_xlsx
from .upload_xlsx import upload_xlsx
from .format_style import format_style

def connect_to_table(show_on_site: dict, pkl_path: str,\
                    xlsx_dict: dict, service_file: str):
    '''
    Connects to the google sheet table, downloads it in excel format, 
        makes changes and uploads it back to the cloud.

    :param show_on_site: a dictionary with the collection statuses of product cards.
    :param pkl_path: a path of pkl file with the product cards.
    :param xlsx_dict: a dictionary with xlsx data from the configuration file.
    :param service_file: a path to google service file.
    :return: zero if success or 1 if error occured.
    '''
    if not os.path.isdir(xlsx_dict['in']):
        Path(xlsx_dict['in']).mkdir(parents=True, exist_ok=True)
    if not os.path.isdir(xlsx_dict['out']):
        Path(xlsx_dict['out']).mkdir(parents=True, exist_ok=True)

    cards = []
    cards_online = []
    cards_offline = []


    if show_on_site['online'] and show_on_site['offline']:
        with open(f"{pkl_path}/online.pkl", 'rb') as file:
            cards_online = pickle.load(file=file)
        with open(f"{pkl_path}/offline.pkl", 'rb') as file:
            cards_offline = pickle.load(file=file)
    elif show_on_site['online']:
        with open(f"{pkl_path}/online.pkl", 'rb') as file:
            cards_online = pickle.load(file=file)
    elif show_on_site['ofline']:
        with open(f"{pkl_path}/offline.pkl", 'rb') as file:
            cards_offline = pickle.load(file=file)
    else:
        raise Exception(f"<connect_to_table> online: {show_on_site['online']}, offline: {show_on_site['offline']}")
    
    alloy_cards = []
    forged_cards = []

    cards = cards_online + cards_offline
    for card in cards:
        if card['type'] == 'alloy':
            alloy_cards.append(card)
        elif card['type'] == 'forged' or card['type'] == 'flow_forming':
            forged_cards.append(card)


    status = download_xlsx(service_file_path=service_file, xlsx_dict=xlsx_dict)

    if status == 0:
        
        file_path_in = os.path.join(xlsx_dict['in'], f"{xlsx_dict['filename']}.xlsx")
        if os.path.exists(file_path_in) and os.path.isfile(file_path_in):
            
            wb = load_workbook(file_path_in)
            
            if not os.path.isdir(xlsx_dict['out']):
                Path(xlsx_dict['out']).mkdir(parents=True, exist_ok=True)
            
            file_path_out = os.path.join(xlsx_dict['out'], f"{xlsx_dict['filename']}.xlsx")
            
            ws_alloy = wb['Литые']
            ws_forged = wb['Кованные']
            
            status = process_xlsx(cards=alloy_cards, xlsx_dict=xlsx_dict, ws=ws_alloy)
            
            if status == 0:
                status = process_xlsx(cards=forged_cards, xlsx_dict=xlsx_dict, ws=ws_forged)
                if status == 0:
                    wb.save(file_path_out)
                    wb.close()

                    status = upload_xlsx(xlsx_dict=xlsx_dict, service_file_path=service_file)
                    if status == 0:
                        status = format_style(xlsx_dict=xlsx_dict, service_file_path=service_file, sheetName='Литые')
                        if status == 0:
                            status = format_style(xlsx_dict=xlsx_dict, service_file_path=service_file, sheetName='Кованные')
                            if status == 0:
                                if 'Для заливки дисков' in wb.sheetnames:
                                    status = format_style(xlsx_dict=xlsx_dict, service_file_path=service_file, sheetName='Для заливки дисков')
                                return 0
        
        
        else:
            raise Exception(f'file: {file_path_in} is not exists')
            

    return 1




